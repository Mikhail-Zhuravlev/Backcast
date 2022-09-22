import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

############################################
# Writes to and saves Excel file
###########################################
def WriteToFile(DFToWrite, TargetFile, TargetSheet, TargetRow, TargetCol):
    
    wb = load_workbook(TargetFile)
    ws = wb.create_sheet(TargetSheet)
    ws.title = TargetSheet
    #ws = wb[TargetSheet]
    
    for rIndex, r in enumerate(dataframe_to_rows(DFToWrite, index=True, header=True)):
        for cIndex, c in enumerate(r):
            ws.cell(row = rIndex + TargetRow, column = cIndex + TargetCol, value = c)

    wb.save(TargetFile)